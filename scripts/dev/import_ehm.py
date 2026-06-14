#!/usr/bin/env python3
"""
DEV-ONLY EHM roster importer.

Reads an EHM "Players and non-players" .xlsx export, filters to the 32 NHL
clubs AND their 32 AHL affiliates, and emits a fictional-DB-shaped mod our
loader understands:

    mods/nhl-ehm/database.json   - ModDatabase (real names/teams/positions)
    mods/nhl-ehm/faces.json      - faceId -> relative faces/<faceId>.png path

Nothing here is committed (mods/ is gitignored). The shipped game stays
fictional; this is only to make local review familiar.

Non-player staff (coaches, scouts, GMs, owners) are collected per NHL club and
emitted as ModTeam.staff[]. The loader builds a real TeamStaff from these
entries, filling missing roles with generated staff. EHM job strings map to our
staff roles as follows:
  "Head Coach"      -> headCoach
  "Assistant Coach" -> assistantCoach
  "General Manager" -> assistantGM   (treated as the GM/decision-maker)
  "Assistant GM"    -> assistantGM
  "Scout"           -> scout
  "Owner"           -> owner
  "Physio"          -> physio
  "Fitness Coach"   -> physio
Rows whose job is "Player" are collected as today; all other job strings that
don't match above are skipped. Staff faces are resolved from the facepack dirs
the same way as player faces (same norm(first_last_dob) key).

AHL affiliates:
  Each NHL ModTeam now carries an optional `affiliate` field (ModAffiliate) with
  the AHL club's city/nickname/abbreviation/colors and its player roster. Players
  whose ClubPlaying matches a known AHL club are routed to that parent's affiliate.
  Only the real AHL players we match are emitted (an affiliate may be thin or
  empty); the game's mod loader tops every affiliate up to valid roster minimums
  with synthesised depth fillers, so we never duplicate an NHL player onto two
  teams.

AHL->NHL parent map (32 affiliates):
  Toronto Marlies     -> TOR    Laval Rocket          -> MTL
  Hershey Bears       -> WSH    Providence Bruins     -> BOS
  Rochester Americans -> BUF    Grand Rapids Griffins -> DET
  Charlotte Checkers  -> FLA    Belleville Senators   -> OTT
  Syracuse Crunch     -> TBL    Cleveland Monsters    -> CBJ
  Utica Comets        -> NJD    Bridgeport Islanders  -> NYI
  Hartford Wolf Pack  -> NYR    Lehigh Valley Phantoms-> PHI
  W-B/Scranton Pens   -> PIT    Rockford IceHogs      -> CHI
  Colorado Eagles     -> COL    Texas Stars           -> DAL
  Iowa Wild           -> MIN    Milwaukee Admirals    -> NSH
  Springfield T-Birds -> STL    Tucson Roadrunners    -> UTA
  Manitoba Moose      -> WPG    San Diego Gulls       -> ANA
  Calgary Wranglers   -> CGY    Bakersfield Condors   -> EDM
  Ontario Reign       -> LAK    San Jose Barracuda    -> SJS
  Coachella Valley    -> SEA    Abbotsford Canucks    -> VAN
  Henderson Silver K. -> VGK    Capitals AHL*         -> WSH (fallback)

* WSH has Hershey Bears as its primary affiliate.

Usage:
    py scripts/dev/import_ehm.py <export.xlsx> <out_dir> [facepack_dir ...]
"""
import sys, os, io, json, unicodedata, glob, shutil

# Column indices (0-based) for the full "Players and non-players" EHM export
# (two header rows: section markers on row 0, field names on row 1; data row 2+).
# Mapped against mods/spreadsheet exports/players_&_non-players.xlsx (133 cols).
C_FIRST, C_SECOND, C_DOB, C_NATION = 1, 2, 3, 4
C_JOB = 9           # "Job For Club" — "Player" for players
C_CLUB = 12         # "Club Playing" — current club (routes AHL/loan players correctly)
C_CA, C_PA = 39, 40
C_GOALIE, C_LD, C_RD, C_LW, C_C, C_RW = 44, 45, 46, 47, 48, 49
C_HAND = 53
C_EXPIRES = 15      # "Contract Expires Club", e.g. "30.6.2027"
C_WAGE = 16         # "Estimated Wage" = annual salary (accurate, e.g. 8700000)
C_ROLE = 52         # "Role", e.g. "Centre: Playmaker", "Defenceman: Defensive"
C_BIRTHTOWN = 7     # "Birth Town", e.g. "cole harbour:ns:can"
C_FAV_NUM, C_SQUAD_NUM = 54, 55   # favourite / squad jersey number
# FM-style personality (each 1-20), maps 1:1 onto our Personality scale.
C_AMBITION, C_DETERMINATION, C_LOYALTY, C_PROFESSIONALISM, C_TEMPERAMENT = 32, 33, 34, 36, 38
C_ADAPTABILITY, C_PRESSURE, C_SPORTSMANSHIP = 31, 35, 37
C_INTL_APPS, C_INTL_GOALS, C_INTL_ASSISTS, C_STANLEY_CUPS = 23, 24, 25, 27
C_HOME_REP, C_CURRENT_REP, C_WORLD_REP = 41, 42, 43
C_JUNIOR_PREF = 59
C_NHL_DRAFT_ELIGIBLE, C_NHL_DRAFTED = 29, 30

# Physical sizes + the full 1-20 attribute columns (EHM scale). Mapped to our
# 1-99 RawAttributes so imported players start out exactly as the DB describes.
C_HEIGHT_CM, C_WEIGHT_KG = 57, 58
C_AGGRESSION, C_ANTICIPATION, C_BRAVERY, C_CONSISTENCY, C_DECISIONS = 60, 61, 62, 63, 64
C_DIRTINESS, C_FLAIR, C_IMPORTANT, C_LEADERSHIP, C_MORALE = 65, 66, 67, 68, 69
C_PASS_TENDENCY, C_TEAMWORK, C_CREATIVITY, C_WORKRATE = 70, 71, 72, 73
C_ACCELERATION, C_AGILITY, C_BALANCE, C_FIGHTING, C_HITTING = 74, 75, 76, 77, 78
C_INJURY_PRONE, C_NAT_FITNESS, C_PACE, C_STAMINA, C_STRENGTH = 79, 80, 81, 82, 83
C_AGITATION, C_CHECKING, C_DEFLECTIONS, C_DEKING, C_FACEOFFS = 84, 85, 86, 87, 88
C_MOVEMENT, C_ONEONONE, C_PASSING, C_POKECHECK, C_POSITIONING = 89, 90, 91, 92, 93
C_SLAPSHOT, C_STICKHANDLING, C_VERSATILITY, C_WRISTSHOT = 94, 95, 96, 97
C_BLOCKER, C_GLOVE, C_REBOUNDS, C_RECOVERY, C_REFLEXES = 98, 99, 100, 101, 102

# Non-player (staff) columns — coaches/scouts/GMs/owners use THESE, not the
# player CA/attribute columns (which are blank for staff).
C_NP_CA = 103            # non-player Current Ability
C_JUDGEMENT, C_JUDGING_POT = 120, 121   # judging ability / potential (1-20)
C_COACH_G, C_COACH_D, C_COACH_F = 116, 117, 118   # coaching goalies/def/fwd
C_COACH_TECH, C_TACTICS_KNOW = 119, 122
C_PHYSIO_SKILL = 123

SEASON_YEAR = 2025

# nickname -> (city, abbr, conference, division, primary, secondary)
NHL = {
    "Bruins": ("Boston", "BOS", "Eastern", "Atlantic", "#FFB81C", "#000000"),
    "Sabres": ("Buffalo", "BUF", "Eastern", "Atlantic", "#003087", "#FFB81C"),
    "Red Wings": ("Detroit", "DET", "Eastern", "Atlantic", "#CE1126", "#FFFFFF"),
    "Panthers": ("Florida", "FLA", "Eastern", "Atlantic", "#041E42", "#C8102E"),
    "Canadiens": ("Montreal", "MTL", "Eastern", "Atlantic", "#AF1E2D", "#192168"),
    "Senators": ("Ottawa", "OTT", "Eastern", "Atlantic", "#C8102E", "#000000"),
    "Lightning": ("Tampa Bay", "TBL", "Eastern", "Atlantic", "#002868", "#FFFFFF"),
    "Maple Leafs": ("Toronto", "TOR", "Eastern", "Atlantic", "#00205B", "#FFFFFF"),
    "Hurricanes": ("Carolina", "CAR", "Eastern", "Metropolitan", "#CC0000", "#000000"),
    "Blue Jackets": ("Columbus", "CBJ", "Eastern", "Metropolitan", "#002654", "#CE1126"),
    "Devils": ("New Jersey", "NJD", "Eastern", "Metropolitan", "#CE1126", "#000000"),
    "Islanders": ("New York", "NYI", "Eastern", "Metropolitan", "#00539B", "#F47D30"),
    "Rangers": ("New York", "NYR", "Eastern", "Metropolitan", "#0038A8", "#CE1126"),
    "Flyers": ("Philadelphia", "PHI", "Eastern", "Metropolitan", "#F74902", "#000000"),
    "Penguins": ("Pittsburgh", "PIT", "Eastern", "Metropolitan", "#FCB514", "#000000"),
    "Capitals": ("Washington", "WSH", "Eastern", "Metropolitan", "#C8102E", "#041E42"),
    "Blackhawks": ("Chicago", "CHI", "Western", "Central", "#CF0A2C", "#000000"),
    "Avalanche": ("Colorado", "COL", "Western", "Central", "#6F263D", "#236192"),
    "Stars": ("Dallas", "DAL", "Western", "Central", "#006847", "#000000"),
    "Wild": ("Minnesota", "MIN", "Western", "Central", "#154734", "#A6192E"),
    "Predators": ("Nashville", "NSH", "Western", "Central", "#FFB81C", "#041E42"),
    "Blues": ("St. Louis", "STL", "Western", "Central", "#002F87", "#FCB514"),
    "Mammoth": ("Utah", "UTA", "Western", "Central", "#69BE28", "#010101"),
    "Jets": ("Winnipeg", "WPG", "Western", "Central", "#041E42", "#004C97"),
    "Ducks": ("Anaheim", "ANA", "Western", "Pacific", "#F47A38", "#B09862"),
    "Flames": ("Calgary", "CGY", "Western", "Pacific", "#C8102E", "#F1BE48"),
    "Oilers": ("Edmonton", "EDM", "Western", "Pacific", "#FF4C00", "#041E42"),
    "Kings": ("Los Angeles", "LAK", "Western", "Pacific", "#111111", "#A2AAAD"),
    "Sharks": ("San Jose", "SJS", "Western", "Pacific", "#006D75", "#000000"),
    "Kraken": ("Seattle", "SEA", "Western", "Pacific", "#001628", "#99D9D9"),
    "Canucks": ("Vancouver", "VAN", "Western", "Pacific", "#00205B", "#00843D"),
    "Golden Knights": ("Vegas", "VGK", "Western", "Pacific", "#B4975A", "#333F42"),
}

# AHL club nickname keywords -> (parent NHL abbreviation, city, full nickname,
#                                 abbreviation, primary, secondary)
# Keyword matching: if any word in this key appears in ClubPlaying (case-insensitive).
AHL_CLUBS = {
    "Marlies":       ("TOR", "Toronto",              "Marlies",       "TOR", "#003E7E", "#FFFFFF"),
    "Rocket":        ("MTL", "Laval",                "Rocket",        "LAV", "#AF1E2D", "#192168"),
    "Hershey":       ("WSH", "Hershey",              "Bears",         "HER", "#862633", "#C0A96A"),
    "Providence":    ("BOS", "Providence",           "Bruins",        "PRO", "#FFB81C", "#000000"),
    "Rochester":     ("BUF", "Rochester",            "Americans",     "ROC", "#003087", "#FFB81C"),
    "Grand Rapids":  ("DET", "Grand Rapids",         "Griffins",      "GRR", "#CE1126", "#FFFFFF"),
    "Charlotte":     ("FLA", "Charlotte",            "Checkers",      "CLT", "#041E42", "#C8102E"),
    "Belleville":    ("OTT", "Belleville",           "Senators",      "BEL", "#C8102E", "#000000"),
    "Syracuse":      ("TBL", "Syracuse",             "Crunch",        "SYR", "#002868", "#FFFFFF"),
    "Cleveland":     ("CBJ", "Cleveland",            "Monsters",      "CLE", "#002654", "#CE1126"),
    "Utica":         ("NJD", "Utica",                "Comets",        "UTI", "#CE1126", "#000000"),
    "Bridgeport":    ("NYI", "Bridgeport",           "Islanders",     "BRI", "#00539B", "#F47D30"),
    "Hartford":      ("NYR", "Hartford",             "Wolf Pack",     "HAR", "#0038A8", "#CE1126"),
    "Lehigh":        ("PHI", "Lehigh Valley",        "Phantoms",      "LHV", "#F74902", "#000000"),
    "Wilkes":        ("PIT", "Wilkes-Barre",         "Penguins",      "WBS", "#FCB514", "#000000"),
    "Rockford":      ("CHI", "Rockford",             "IceHogs",       "RFD", "#CF0A2C", "#000000"),
    "Colorado Eagles": ("COL", "Colorado",           "Eagles",        "COE", "#6F263D", "#236192"),
    "Texas":         ("DAL", "Cedar Park",           "Stars",         "TEX", "#006847", "#000000"),
    "Iowa":          ("MIN", "Des Moines",           "Wild",          "IOW", "#154734", "#A6192E"),
    "Milwaukee":     ("NSH", "Milwaukee",            "Admirals",      "MIL", "#FFB81C", "#041E42"),
    "Springfield":   ("STL", "Springfield",          "Thunderbirds",  "SPR", "#002F87", "#FCB514"),
    "Tucson":        ("UTA", "Tucson",               "Roadrunners",   "TUC", "#69BE28", "#010101"),
    "Manitoba":      ("WPG", "Winnipeg",             "Moose",         "MAN", "#041E42", "#004C97"),
    "San Diego":     ("ANA", "San Diego",            "Gulls",         "SDG", "#F47A38", "#B09862"),
    "Wranglers":     ("CGY", "Calgary",              "Wranglers",     "CGW", "#C8102E", "#F1BE48"),
    "Bakersfield":   ("EDM", "Bakersfield",          "Condors",       "BAK", "#FF4C00", "#041E42"),
    "Ontario":       ("LAK", "Ontario",              "Reign",         "ONT", "#111111", "#A2AAAD"),
    "Barracuda":     ("SJS", "San Jose",             "Barracuda",     "SJB", "#006D75", "#000000"),
    "Coachella":     ("SEA", "Palm Desert",          "Firebirds",     "CVF", "#001628", "#99D9D9"),
    "Abbotsford":    ("VAN", "Abbotsford",           "Canucks",       "ABB", "#00205B", "#00843D"),
    "Henderson":     ("VGK", "Henderson",            "Silver Knights","HEN", "#B4975A", "#333F42"),
}

def norm(s):
    """Lowercase, strip accents, spaces->underscores — for face matching."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return s.lower().replace(" ", "_").strip()

def match_nhl_club(club):
    """Return the NHL nickname key if the club string matches an NHL team."""
    if not club or club == "[None]":
        return None
    for nick in NHL:
        if nick.lower() in str(club).lower():
            return nick
    return None

def match_ahl_club(club):
    """Return the AHL keyword key if the club string matches a known AHL affiliate."""
    if not club or club == "[None]":
        return None
    club_lower = str(club).lower()
    for keyword in AHL_CLUBS:
        # Multi-word keywords (e.g. "Grand Rapids", "San Diego") need all words present.
        if all(word.lower() in club_lower for word in keyword.split()):
            return keyword
    return None

def map_role(role_str, pos):
    """Map an EHM role string (e.g. 'Centre: Playmaker') to our PlayerRole."""
    s = str(role_str or "").lower()
    if pos == "G":
        return "starter"
    if pos == "D":
        if "offensive" in s or "puck" in s or "two" in s:
            return "offensiveD"
        if "stay" in s:
            return "stayAtHomeD"
        return "shutdownD"  # "Defensive", "All around", default
    # Forwards.
    if "playmaker" in s:
        return "playmaker"
    if "sniper" in s or "goal" in s or "scorer" in s:
        return "sniper"
    if "power" in s:
        return "powerForward"
    if "enforcer" in s or "fighter" in s:
        return "enforcer"
    return "twoWay"  # "Two-way", "Defensive", "Grinder", "All around", default

def contract_from_row(row):
    """Build {salary, years} from the EHM wage + contract-expiry columns."""
    salary = to_int(row[C_WAGE], 0)
    if salary <= 0:
        return None
    parts = str(row[C_EXPIRES] or "").split(".")
    exp_year = to_int(parts[2]) if len(parts) == 3 else 0
    years = clamp(exp_year - SEASON_YEAR, 1, 8) if exp_year else 2
    return {"salary": salary, "years": years}

def personality_from_row(row):
    """EHM personality columns (1-20) -> our Personality (same 1-20 scale)."""
    def p(col):
        return clamp(to_int(row[col], 10), 1, 20)
    return {
        "ambition": p(C_AMBITION),
        "professionalism": p(C_PROFESSIONALISM),
        "loyalty": p(C_LOYALTY),
        "temperament": p(C_TEMPERAMENT),
        "determination": p(C_DETERMINATION),
    }

def extended_fields_from_row(row):
    """Build the extended EHM-sourced fields for a player dict.
    Only emits keys that are meaningful (non-zero / non-empty).
    All returned keys are optional; absence is always a valid fallback."""
    out = {}

    # Gameplay attrs (EHM 1-20 -> our 1-99 via s20)
    for key, col in [
        ("leadership",    C_LEADERSHIP),
        ("teamwork",      C_TEAMWORK),
        ("flair",         C_FLAIR),
        ("agitation",     C_AGITATION),
        ("movement",      C_MOVEMENT),
        ("oneOnOnes",     C_ONEONONE),
        ("versatility",   C_VERSATILITY),
    ]:
        raw = to_int(row[col], 0)
        if raw > 0:
            out[key] = s20(row, col)

    # fighting is also 1-20 -> 1-99
    raw_fight = to_int(row[C_FIGHTING], 0)
    if raw_fight > 0:
        out["fighting"] = s20(row, C_FIGHTING)

    # injuryProneness / naturalFitness (1-20 -> 1-99)
    raw_ip = to_int(row[C_INJURY_PRONE], 0)
    if raw_ip > 0:
        out["injuryProneness"] = s20(row, C_INJURY_PRONE)
    raw_nf = to_int(row[C_NAT_FITNESS], 0)
    if raw_nf > 0:
        out["naturalFitness"] = s20(row, C_NAT_FITNESS)

    # Personality-adjacent (keep EHM 1-20 scale, clamp 1-20)
    for key, col in [
        ("adaptability",  C_ADAPTABILITY),
        ("pressure",      C_PRESSURE),
        ("sportsmanship", C_SPORTSMANSHIP),
    ]:
        raw = to_int(row[col], 0)
        if raw > 0:
            out[key] = clamp(raw, 1, 20)

    # History counts (emit only if > 0)
    for key, col in [
        ("intlApps",    C_INTL_APPS),
        ("intlGoals",   C_INTL_GOALS),
        ("intlAssists", C_INTL_ASSISTS),
        ("stanleyCups", C_STANLEY_CUPS),
    ]:
        val = to_int(row[col], 0)
        if val > 0:
            out[key] = val

    # Reputation (0-200, emit as-is; emit if > 0)
    for key, col in [
        ("homeReputation",    C_HOME_REP),
        ("currentReputation", C_CURRENT_REP),
        ("worldReputation",   C_WORLD_REP),
    ]:
        val = to_int(row[col], 0)
        if val > 0:
            out[key] = val

    # Draft booleans — coerce from Python True/False or string 'True'/'False'
    def to_bool(v):
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in ("true", "1", "yes"):
            return True
        if s in ("false", "0", "no", "[none]", ""):
            return False
        return None

    eligible = to_bool(row[C_NHL_DRAFT_ELIGIBLE] if len(row) > C_NHL_DRAFT_ELIGIBLE else None)
    if eligible is not None:
        out["nhlDraftEligible"] = eligible

    drafted = to_bool(row[C_NHL_DRAFTED] if len(row) > C_NHL_DRAFTED else None)
    if drafted is not None:
        out["nhlDrafted"] = drafted

    # Junior preference string (skip '[None]' and empty)
    jp = str(row[C_JUNIOR_PREF] or "").strip()
    if jp and jp != "[None]":
        out["juniorPreference"] = jp

    return out


def birthplace_from_row(row):
    """EHM birth town 'cole harbour:ns:can' -> 'Cole Harbour, NS'."""
    raw = str(row[C_BIRTHTOWN] or "").strip()
    if not raw or raw == "[None]":
        return None
    parts = [x for x in raw.split(":") if x]
    if not parts:
        return None
    town = parts[0].title()
    if len(parts) >= 2 and len(parts[1]) <= 3:
        return f"{town}, {parts[1].upper()}"
    return town

def position(row):
    g = row[C_GOALIE] or 0
    d = max(row[C_LD] or 0, row[C_RD] or 0)
    c = row[C_C] or 0
    w = max(row[C_LW] or 0, row[C_RW] or 0)
    best = max((g, "G"), (d, "D"), (c, "C"), (w, "W"), key=lambda t: t[0])
    return best[1]

def clamp(v, lo, hi):
    return max(lo, min(hi, v))

def to_int(v, default=0):
    try:
        return int(float(v))
    except Exception:
        return default

def s20(row, col):
    """EHM 1-20 attribute -> our 1-99 scale."""
    return clamp(round(to_int(row[col], 0) / 20.0 * 99), 1, 99)

def resolve_potential(pa, ca, age):
    """Return (single_ca, band_or_None) on the 1-200 CA scale.

    Positive PA is an explicit ceiling, used as-is. Negative PA is an OPAQUE
    range code in this DB (it spans -1..-20 and its magnitude is not a reliable
    'higher = better' signal — the most common codes are mid players). So for
    negatives we derive a realistic ceiling from CURRENT ability + youth
    headroom: a player only projects as a star if he's already strong for his
    age. The engine rolls within the [lo,hi] band per career (option B)."""
    if pa is None or pa >= 0:
        return max(pa or ca, ca), None
    base = (46 if age <= 18 else 38 if age == 19 else 30 if age == 20 else
            23 if age == 21 else 16 if age == 22 else 11 if age == 23 else 7)
    lo = min(200, ca + round(base * 0.55))
    hi = min(200, ca + round(base * 1.15))
    lo = max(lo, ca); hi = max(hi, lo)
    return (lo + hi) // 2, (lo, hi)

def avg20(row, *cols):
    """Average of several EHM 1-20 attributes -> our 1-99 scale."""
    vals = [to_int(row[c], 0) for c in cols]
    return clamp(round(sum(vals) / len(vals) / 20.0 * 99), 1, 99)

def height_rating(row):
    """Height in cm -> a 1-99 size rating (≈165cm -> 1, ≈203cm -> 99)."""
    cm = to_int(row[C_HEIGHT_CM], 0)
    if cm <= 0:
        return None
    return clamp(round((cm - 165) / (203 - 165) * 99), 1, 99)

def build_attributes(row, pos):
    """Map the EHM 1-20 attribute columns onto our flat ModPlayerAttributes
    (1-99 each), so an imported player starts out exactly as the DB describes.
    Attributes our model has but EHM doesn't are derived from the closest EHM
    inputs (e.g. offensiveIQ from Decisions+Creativity)."""
    a = {
        # Technical
        "wristShot": s20(row, C_WRISTSHOT),
        "slapShot": s20(row, C_SLAPSHOT),
        "stickhandling": avg20(row, C_STICKHANDLING, C_DEKING),
        "passing": s20(row, C_PASSING),
        "deflections": s20(row, C_DEFLECTIONS),
        "faceoffs": s20(row, C_FACEOFFS),
        # Physical
        "speed": s20(row, C_PACE),
        "acceleration": s20(row, C_ACCELERATION),
        "strength": s20(row, C_STRENGTH),
        "balance": s20(row, C_BALANCE),
        "stamina": s20(row, C_STAMINA),
        "agility": s20(row, C_AGILITY),
        # Mental
        "offensiveIQ": avg20(row, C_DECISIONS, C_CREATIVITY),
        "defensiveIQ": avg20(row, C_DECISIONS, C_POSITIONING),
        "positioning": s20(row, C_POSITIONING),
        "vision": s20(row, C_CREATIVITY),
        "aggression": s20(row, C_AGGRESSION),
        "composure": avg20(row, C_CONSISTENCY, C_IMPORTANT),
        "workRate": s20(row, C_WORKRATE),
        # Dirtiness is inverse of discipline.
        "discipline": clamp(round((20 - to_int(row[C_DIRTINESS], 5)) / 20.0 * 99), 1, 99),
        "anticipation": s20(row, C_ANTICIPATION),
        # Defensive
        "checking": s20(row, C_CHECKING),
        "shotBlocking": avg20(row, C_BRAVERY, C_CHECKING),
        "stickChecking": s20(row, C_POKECHECK),
        "takeaway": avg20(row, C_POKECHECK, C_ANTICIPATION),
    }
    h = height_rating(row)
    if h is not None:
        a["height"] = h
    if pos == "G":
        a.update({
            "reflexes": s20(row, C_REFLEXES),
            "positioningG": s20(row, C_POSITIONING),
            "reboundControl": s20(row, C_REBOUNDS),
            "glove": s20(row, C_GLOVE),
            "blocker": s20(row, C_BLOCKER),
            "recovery": s20(row, C_RECOVERY),
            "puckHandlingG": avg20(row, C_STICKHANDLING, C_PASSING),
        })
    return a

def map_staff_role(job_str):
    """Map an EHM job string to our StaffMember role union, or return None to skip.

    EHM job strings are title-case, e.g. 'Head Coach', 'Assistant Coach',
    'General Manager', 'Assistant General Manager', 'Scout', 'Owner',
    'Physio', 'Fitness Coach'. Rows whose job is 'Player' are handled
    separately in the player loop and must NOT be passed here.
    """
    s = str(job_str or "").lower()
    if "head coach" in s:
        return "headCoach"
    if "assistant coach" in s or s.strip() == "coach" or "goaltending coach" in s or "skills coach" in s:
        return "assistantCoach"
    if "assistant" in s and ("gm" in s or "general manager" in s):
        return "assistantGM"
    if "general manager" in s:
        return None  # the GM is the human player's job — don't add as staff
    if "chairman" in s or "owner" in s or "governor" in s or "president" in s:
        return "owner"
    if "scout" in s or "director of personnel" in s:
        return "scout"
    if "physio" in s or "fitness" in s or "trainer" in s or "therapist" in s:
        return "physio"
    return None  # skip all other front-office roles


# Non-player attribute columns (EHM 1-20) -> our StaffAttributes keys.
STAFF_ATTR_COLS = {
    "attacking": 109, "directness": 110, "freeRoles": 111, "lineMatching": 112,
    "penaltyKill": 113, "physical": 114, "powerplay": 115,
    "coachingGoaltenders": 116, "coachingDefensemen": 117, "coachingForwards": 118,
    "coachingTechnique": 119, "judgingPlayers": 120, "judgingPotential": 121,
    "tactics": 122, "physiotherapy": 123, "business": 124, "patience": 125,
    "resources": 126, "discipline": 127, "manManagement": 129, "motivating": 130,
    "developingYoungsters": 131,
}

def staff_attrs(r):
    """Per-discipline staff attributes (1-20), omitting zeros."""
    out = {}
    for key, col in STAFF_ATTR_COLS.items():
        v = to_int(r[col], 0)
        if v > 0:
            out[key] = v
    return out


def staff_specialty(r, role):
    """Specialty from the EHM non-player coaching/judging columns."""
    if role in ("headCoach", "assistantCoach"):
        opts = [
            (to_int(r[C_COACH_F], 0), "Forwards"),
            (to_int(r[C_COACH_D], 0), "Defense"),
            (to_int(r[C_COACH_G], 0), "Goaltending"),
            (to_int(r[C_TACTICS_KNOW], 0), "Tactics"),
        ]
        opts.sort(key=lambda t: -t[0])
        return opts[0][1] if opts[0][0] > 0 else None
    if role == "scout":
        return "Prospects" if to_int(r[C_JUDGING_POT], 0) >= to_int(r[C_JUDGEMENT], 0) else "Pro Scouting"
    if role == "physio":
        return "Fitness"
    return None


# player_career_history.xlsx column indices (positional).
CH_FIRST, CH_SECOND, CH_DOB = 1, 2, 3
CH_ONLOAN, CH_PLAYOFFS, CH_YEAR, CH_CLUB, CH_COMP = 4, 5, 6, 7, 8
CH_GP, CH_G, CH_A, CH_PIM, CH_PM = 9, 10, 11, 12, 13
CH_MINS, CH_GA, CH_SO, CH_W, CH_L, CH_TOT, CH_SAVES = 14, 15, 16, 17, 18, 19, 20


def _sibling(input_xlsx, name, env):
    e = os.environ.get(env)
    if e and os.path.exists(e):
        return e
    sib = os.path.join(os.path.dirname(os.path.abspath(input_xlsx)), name)
    if os.path.exists(sib):
        return sib
    repo = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    guess = os.path.join(repo, "mods", "spreadsheet exports", name)
    return guess if os.path.exists(guess) else None


def load_draft_history(path, keep_keys):
    """draft_history.xlsx -> key -> {year, round, overall, club}. Keyed name+DOB.
    Cols: Draft1 Year2 Round3 Overall4 Club5 First6 Second7 DOB8."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it)
    out = {}
    for r in it:
        first = str(r[6] or "").strip(); last = str(r[7] or "").strip()
        parts = str(r[8] or "").split(".")
        if len(parts) != 3:
            continue
        key = norm(f"{first}_{last}_{'_'.join(parts)}")
        if key not in keep_keys or key in out:
            continue
        out[key] = {
            "draftYear": to_int(r[2], 0),
            "draftRound": to_int(r[3], 0),
            "draftOverall": to_int(r[4], 0),
            "draftClub": str(r[5] or "").strip(),
        }
    wb.close()
    return out


def load_clubs(path):
    """clubs.xlsx (1 header row) -> full-name(lower) -> {arena, capacity} for NHL clubs.
    Cols: Name1 Abbreviation4 Division5 Arena13 MaxAttendance19 Cash20 PlayerBudget21."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it)
    out = {}
    for r in it:
        if str(r[5] or "").strip() != "National Hockey League":
            continue
        name = str(r[1] or "").strip()
        if not name:
            continue
        out[name.lower()] = {
            "arena": str(r[13] or "").strip(),
            "capacity": to_int(r[19], 0),
        }
    wb.close()
    return out


# ── Multi-league world (#95) ────────────────────────────────────────────
# Top-level league names (exact, from club_competitions / clubs "Division")
# to import as quick-sim/background competitions, mapped to the abbrev the
# engine's NHLe leagueStrength model recognises. NCAA is split into conferences
# in the DB; each is imported as its own competition tagged 'NCAA' for strength.
COMP_LEAGUES = {
    "Ontario Hockey League": "OHL",
    "Western Hockey League": "WHL",
    "United States Hockey League": "USHL",
    "Kontinental Hockey League": "KHL",
    "Swedish Hockey League": "SHL",
    "Finnish Liiga": "LIIGA",
    "Swiss National League": "NL",
    "Czech Tipsport Extraliga": "EXTRALIGA",
    "Deutsche Eishockey Liga": "DEL",
    "ECHL": "ECHL",
    "Swedish HockeyAllsvenskan": "HA",
    # NCAA Division I — clubs file under this umbrella name (where US college
    # draft prospects like Gavin McKenna play).
    "National Collegiate Athletic Association": "NCAA",
}
COMP_BY_LOWER = {k.lower(): k for k in COMP_LEAGUES}

# Deterministic team color palette ('#RRGGBB' primary, secondary) for imported
# competition clubs (the source DB has no per-club colors).
COMP_PALETTE = [
    ("#1f4e8c", "#c8d6e5"), ("#8c1f2f", "#e5c8cc"), ("#1f8c5a", "#c8e5d6"),
    ("#8c6a1f", "#e5dcc8"), ("#4a1f8c", "#d6c8e5"), ("#1f7d8c", "#c8e1e5"),
    ("#8c3d1f", "#e5d2c8"), ("#5a8c1f", "#d8e5c8"), ("#2b2f38", "#cfd4dc"),
    ("#8c1f6a", "#e5c8dc"),
]


def match_comp_league(div):
    """Canonical whitelisted league name for a club's Division (exact match), or
    None. Exact-only avoids pulling in sub-tiers like 'NCAA … Division III'."""
    return COMP_BY_LOWER.get(str(div or "").strip().lower())


def load_competition_meta(path):
    """club_competitions.xlsx -> name(lower) -> {abbrev, nation, level, reputation, ageLimit}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True); next(it)
    out = {}
    for r in it:
        name = str(r[2] or "").strip()
        if not name:
            continue
        out[name.lower()] = {
            "abbrev": str(r[4] or "").strip(),
            "nation": str(r[6] or "").strip(),
            "level": to_int(r[8], 1),
            "reputation": to_int(r[10], 10),
            "ageLimit": to_int(r[11], 0),
        }
    wb.close()
    return out


def load_club_league_map(path):
    """clubs.xlsx -> club-name(lower) -> {name, nickname, abbr, city, nation, league}
    for clubs whose Division is a whitelisted competition league."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True); next(it)
    out = {}
    for r in it:
        league = match_comp_league(r[5])
        if league is None:
            continue
        name = str(r[1] or "").strip()
        if not name:
            continue
        # HomeCity is "city:region:country" — take the city segment, title-cased.
        raw_city = str(r[11] or "").strip().split(":")[0].replace("_", " ").strip()
        out[name.lower()] = {
            "name": name,
            "nickname": str(r[3] or "").strip() or name.split()[-1],
            "abbr": (str(r[4] or "").strip() or name[:3]).upper()[:3],
            "city": raw_city.title(),
            "nation": str(r[12] or "").strip(),
            "league": league,
        }
    wb.close()
    return out


def load_retired_numbers(path):
    """retired_numbers.xlsx (2 header rows) -> full-club-name(lower) -> [{number, player}]."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it); next(it)
    out = {}
    for r in it:
        club = str(r[0] or "").strip()
        num = to_int(r[1], 0)
        player = str(r[2] or "").strip()
        if not club or num <= 0:
            continue
        out.setdefault(club.lower(), []).append({"number": num, "player": player})
    wb.close()
    return out


def find_career_history(input_xlsx):
    """Locate player_career_history.xlsx: sibling of the input, the repo's
    'mods/spreadsheet exports' folder, or the EHM_CAREER_HISTORY env override."""
    env = os.environ.get("EHM_CAREER_HISTORY")
    if env and os.path.exists(env):
        return env
    sibling = os.path.join(os.path.dirname(os.path.abspath(input_xlsx)), "player_career_history.xlsx")
    if os.path.exists(sibling):
        return sibling
    repo = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    guess = os.path.join(repo, "mods", "spreadsheet exports", "player_career_history.xlsx")
    return guess if os.path.exists(guess) else None


def load_career_history(path, keep_keys):
    """Stream the (large) career-history sheet, keeping regular-season rows only
    for players in keep_keys (norm first_last_dob). Returns key -> [season dicts]."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it)  # header
    out = {}
    for r in it:
        if str(r[CH_PLAYOFFS] or "").strip().lower() == "y":
            continue  # regular season only
        first = str(r[CH_FIRST] or "").strip()
        last = str(r[CH_SECOND] or "").strip()
        dob = str(r[CH_DOB] or "")
        parts = dob.split(".")
        if len(parts) != 3:
            continue
        key = norm(f"{first}_{last}_{'_'.join(parts)}")
        if key not in keep_keys:
            continue
        gp = to_int(r[CH_GP], 0)
        if gp <= 0:
            continue
        out.setdefault(key, []).append({
            "year": to_int(r[CH_YEAR], 0),
            "club": str(r[CH_CLUB] or "").strip(),
            "league": str(r[CH_COMP] or "").strip(),
            "gp": gp,
            "g": to_int(r[CH_G], 0),
            "a": to_int(r[CH_A], 0),
            "pim": to_int(r[CH_PIM], 0),
            "plusMinus": to_int(r[CH_PM], 0),
            "mins": to_int(r[CH_MINS], 0),
            "ga": to_int(r[CH_GA], 0),
            "so": to_int(r[CH_SO], 0),
            "w": to_int(r[CH_W], 0),
            "l": to_int(r[CH_L], 0),
            "otl": to_int(r[CH_TOT], 0),
            "saves": to_int(r[CH_SAVES], 0),
        })
    wb.close()
    # Newest first, cap to a sane number of seasons per player.
    for k in out:
        out[k].sort(key=lambda s: -s["year"])
        del out[k][25:]
    return out


def main():
    xlsx, out_dir = sys.argv[1], sys.argv[2]
    facedirs = sys.argv[3:]
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    it = ws.iter_rows(values_only=True); next(it); next(it)

    # team -> list of player dicts
    nhl_teams = {nick: [] for nick in NHL}
    # ahl_keyword -> list of player dicts
    ahl_teams = {kw: [] for kw in AHL_CLUBS}
    # NHL nickname -> list of staff dicts (ModStaff shape)
    nhl_staff = {nick: [] for nick in NHL}

    # Multi-league world: load club->league + league metadata up front so the
    # player loop can route non-NHL/AHL players into their competition club.
    _clubs_path = _sibling(xlsx, "clubs.xlsx", "EHM_CLUBS")
    _comps_path = _sibling(xlsx, "club_competitions.xlsx", "EHM_CLUB_COMPETITIONS")
    club_league_map = load_club_league_map(_clubs_path) if _clubs_path else {}
    comp_meta = load_competition_meta(_comps_path) if _comps_path else {}
    # league(proper) -> club-name -> {"info": club_info, "players": [...]}
    comp_clubs = {}

    for r in it:
        job = r[C_JOB]
        job_str = str(job or "")
        club = r[C_CLUB]
        is_player = job and "player" in job_str.lower()

        if not is_player:
            # Non-player: attempt to collect as staff for an NHL club.
            staff_role = map_staff_role(job_str)
            if staff_role is not None:
                nhl_nick = match_nhl_club(club)
                if nhl_nick is not None:
                    first = str(r[C_FIRST] or "").strip()
                    last = str(r[C_SECOND] or "").strip()
                    if first or last:
                        dob = str(r[C_DOB] or "")
                        parts = dob.split(".")
                        face_id = norm(f"{first}_{last}_{'_'.join(parts)}") if len(parts) == 3 else None
                        # Staff use the NON-PLAYER ability column (player CA is blank).
                        np_ca = to_int(r[C_NP_CA], 0)
                        rating = clamp(round(np_ca / 2), 1, 99) if np_ca > 0 else 40
                        # Judgment from the EHM judging-ability / judging-potential (1-20).
                        jraw = max(to_int(r[C_JUDGEMENT], 0), to_int(r[C_JUDGING_POT], 0))
                        judgment = clamp(round(jraw / 20.0 * 100), 0, 100) if jraw > 0 else None
                        specialty = staff_specialty(r, staff_role)
                        staff_obj = {
                            "name": f"{first} {last}".strip(),
                            "role": staff_role,
                            "rating": rating,
                        }
                        if judgment is not None:
                            staff_obj["judgment"] = judgment
                        if specialty is not None:
                            staff_obj["specialty"] = specialty
                        attrs = staff_attrs(r)
                        if attrs:
                            staff_obj["attributes"] = attrs
                        if face_id is not None:
                            staff_obj["_faceId"] = face_id  # resolved below
                        nhl_staff[nhl_nick].append(staff_obj)
            continue

        # ── Players (original logic unchanged below) ─────────────────────

        # Match AHL FIRST: many AHL club names contain an NHL nickname
        # ("Providence Bruins", "Belleville Senators", "Wilkes-Barre Penguins",
        # "Abbotsford Canucks", "Bridgeport Islanders"), so an NHL-first check
        # would wrongly route those affiliate players into the NHL pool.
        ahl_kw = match_ahl_club(club)
        nhl_nick = None if ahl_kw else match_nhl_club(club)

        # Route non-NHL/AHL players into a whitelisted competition club, if any.
        comp_club_info = None
        if nhl_nick is None and ahl_kw is None:
            comp_club_info = club_league_map.get(str(club or "").strip().lower())
            if comp_club_info is None:
                continue

        first, last = str(r[C_FIRST] or "").strip(), str(r[C_SECOND] or "").strip()
        if not first and not last:
            continue
        dob = str(r[C_DOB] or "")
        parts = dob.split(".")
        year = to_int(parts[2]) if len(parts) == 3 else 0
        age = clamp(SEASON_YEAR - year if year else 25, 16, 45)
        ca = to_int(r[C_CA], 50)
        pa_single, pa_band = resolve_potential(to_int(r[C_PA], ca), ca, age)
        face_id = norm(f"{first}_{last}_{'_'.join(parts)}") if len(parts) == 3 else None
        pos = position(r)
        extended = extended_fields_from_row(r)
        player = {
            "externalId": f"ehm-{norm(first)}-{norm(last)}-{year}",
            "name": f"{first} {last}".strip(),
            "age": age,
            "position": pos,
            "handedness": "L" if str(r[C_HAND] or "Right").lower().startswith("l") else "R",
            # overall is a caliber fallback only; the loader derives the shown
            # overall from the real attributes below. potential (PA) is the DB's
            # potential-ability number (negative = range code), halved to 1-99.
            "overall": clamp(round(ca / 2), 1, 99),
            "potential": clamp(round(pa_single / 2), 1, 99),
            **({"potentialRange": [clamp(round(pa_band[0] / 2), 1, 99), clamp(round(pa_band[1] / 2), 1, 99)]} if pa_band else {}),
            "attributes": build_attributes(r, pos),
            "role": map_role(r[C_ROLE], pos),
            "personality": personality_from_row(r),
            "faceId": face_id,
            "_ca": ca,
            "_pa": pa_single,
            "_key": face_id,  # norm(first_last_dob) — joins to player_career_history
            **extended,
        }
        contract = contract_from_row(r)
        if contract is not None:
            player["contract"] = contract
        nationality = str(r[C_NATION] or "").strip()
        if nationality and nationality != "[None]":
            player["nationality"] = nationality
        bp = birthplace_from_row(r)
        if bp is not None:
            player["birthplace"] = bp
        jersey = to_int(r[C_SQUAD_NUM], 0) or to_int(r[C_FAV_NUM], 0)
        if jersey > 0:
            player["jerseyNumber"] = jersey
        ht = to_int(r[C_HEIGHT_CM], 0)
        if ht > 0:
            player["heightCm"] = ht
        wt = to_int(r[C_WEIGHT_KG], 0)
        if wt > 0:
            player["weightKg"] = wt
        if nhl_nick:
            nhl_teams[nhl_nick].append(player)
        elif ahl_kw:
            ahl_teams[ahl_kw].append(player)
        else:
            league = comp_club_info["league"]
            cname = comp_club_info["name"]
            bucket = comp_clubs.setdefault(league, {}).setdefault(
                cname, {"info": comp_club_info, "players": []})
            bucket["players"].append(player)

    # Attach real season-by-season career history from the DB export, for the
    # players we actually kept (NHL rosters + AHL affiliates + overflow).
    all_players = [p for plist in nhl_teams.values() for p in plist] + \
                  [p for plist in ahl_teams.values() for p in plist] + \
                  [p for clubs in comp_clubs.values() for b in clubs.values() for p in b["players"]]
    keep_keys = {p["_key"] for p in all_players if p.get("_key")}
    hist_path = find_career_history(xlsx)
    if hist_path:
        print(f"  career history: {hist_path}")
        career = load_career_history(hist_path, keep_keys)
        attached = 0
        for p in all_players:
            k = p.get("_key")
            if k and k in career:
                p["careerHistory"] = career[k]
                attached += 1
        print(f"  career history attached to {attached}/{len(all_players)} players")
        # Draft history (which club drafted each player, year/round/overall).
        draft_path = _sibling(xlsx, "draft_history.xlsx", "EHM_DRAFT_HISTORY")
        if draft_path:
            draft = load_draft_history(draft_path, keep_keys)
            dn = 0
            for p in all_players:
                k = p.get("_key")
                if k and k in draft:
                    p.update(draft[k]); dn += 1
            print(f"  draft history attached to {dn}/{len(all_players)} players")
    else:
        print("  career history: not found (skipping) — set EHM_CAREER_HISTORY or place player_career_history.xlsx beside the export")

    # Build face index from the facepack dirs (filename without .png -> path).
    face_index = {}
    for d in facedirs:
        for p in glob.glob(os.path.join(d, "*.png")):
            face_index[norm(os.path.splitext(os.path.basename(p))[0])] = p

    # Copy matched faces into the mod's own faces/ folder (self-contained,
    # per MODDING.md: faces/<faceId>.png). Keeps the mod portable instead of
    # depending on the multi-GB external facepacks at runtime.
    faces_dir = os.path.join(out_dir, "faces")
    os.makedirs(faces_dir, exist_ok=True)

    def resolve_faces(players, faces_out, total_ref, matched_ref):
        """Copy face images for a player list, return cleaned player list."""
        clean = []
        for p in players:
            total_ref[0] += 1
            fid = p.get("faceId")
            if fid and fid in face_index:
                dest = os.path.join(faces_dir, fid + ".png")
                if not os.path.exists(dest):
                    try:
                        shutil.copyfile(face_index[fid], dest)
                    except Exception:
                        pass
                faces_out[fid] = fid + ".png"
                matched_ref[0] += 1
            q = {k: v for k, v in p.items() if not k.startswith("_")}
            clean.append(q)
        return clean

    def resolve_staff_faces(staff_list, faces_out, staff_total_ref, staff_matched_ref):
        """Resolve faceId for staff entries and clean _faceId sentinel."""
        clean = []
        for s in staff_list:
            staff_total_ref[0] += 1
            fid = s.get("_faceId")
            out = {k: v for k, v in s.items() if k != "_faceId"}
            if fid and fid in face_index:
                dest = os.path.join(faces_dir, fid + ".png")
                if not os.path.exists(dest):
                    try:
                        shutil.copyfile(face_index[fid], dest)
                    except Exception:
                        pass
                faces_out[fid] = fid + ".png"
                out["faceId"] = fid
                staff_matched_ref[0] += 1
            clean.append(out)
        return clean

    # Build reverse lookup: NHL abbreviation -> NHL nickname key.
    abbr_to_nick = {v[1]: k for k, v in NHL.items()}

    # Club arena/capacity + retired numbers from the DB (matched by full name
    # ending in our NHL nickname, e.g. "Pittsburgh Penguins" -> "Penguins").
    clubs_path = _sibling(xlsx, "clubs.xlsx", "EHM_CLUBS")
    retired_path = _sibling(xlsx, "retired_numbers.xlsx", "EHM_RETIRED")
    clubs = load_clubs(clubs_path) if clubs_path else {}
    retired = load_retired_numbers(retired_path) if retired_path else {}
    def club_meta(nick):
        nl = nick.lower()
        for name, meta in clubs.items():
            if name.endswith(nl):
                return meta
        return None
    def club_retired(nick):
        nl = nick.lower()
        for name, lst in retired.items():
            if name.endswith(nl):
                return lst
        return None

    # Assemble conferences/divisions/teams; per NHL team take a sensible roster.
    confs = {}
    faces_out = {}
    total, matched = [0], [0]
    ahl_total, ahl_matched = [0], [0]
    staff_total, staff_matched = [0], [0]

    for nick, (city, abbr, conf, div, prim, sec) in NHL.items():
        roster = sorted(nhl_teams[nick], key=lambda p: -p["_ca"])
        goalies = [p for p in roster if p["position"] == "G"][:3]
        d = [p for p in roster if p["position"] == "D"][:8]
        fwd = [p for p in roster if p["position"] in ("C", "W")][:14]
        chosen = goalies[:max(2, len(goalies))] + d + fwd
        overflow = [p for p in roster if p not in chosen]  # NHL players not selected

        if len(goalies) < 2 or len(d) < 7 or len(fwd) < 13:
            print(f"  WARN NHL {nick}: G{len(goalies)} D{len(d)} F{len(fwd)}")

        clean_nhl = resolve_faces(chosen, faces_out, total, matched)

        # Build the affiliate section for this NHL team.
        # Find the AHL club whose parent maps to this NHL abbreviation.
        aff_kw = None
        for kw, (parent_abbr, *_) in AHL_CLUBS.items():
            if parent_abbr == abbr:
                aff_kw = kw
                break

        affiliate_obj = None
        if aff_kw is not None:
            aff_info = AHL_CLUBS[aff_kw]  # (parent_abbr, city, nickname, abbr3, prim, sec)
            _, aff_city, aff_nick, aff_abbr, aff_prim, aff_sec = aff_info
            # Affiliate pool = players whose club matched the AHL keyword PLUS the
            # NHL org's overflow (players beyond the 23-man NHL roster). Folding
            # overflow in keeps real org prospects (e.g. recently-signed kids on
            # the NHL club's reserve list) instead of dropping them on the floor.
            # Dedupe by identity so a player is never on both NHL and AHL rosters.
            aff_pool = list(ahl_teams[aff_kw]) + list(overflow)
            # Prefer keeping young, high-potential players: rank by a blend of
            # current ability and upside (younger + higher PA scores higher) so
            # the caps below cut aging depth rather than prospects.
            def _aff_rank(p):
                ca = p.get("_ca", 0)
                pa = p.get("_pa", ca)
                age = p.get("age", 27)
                youth_bonus = max(0, 24 - age) * 1.5
                return -(ca * 0.5 + pa * 0.5 + youth_bonus)
            aff_roster = sorted(aff_pool, key=_aff_rank)
            aff_goalies = [p for p in aff_roster if p["position"] == "G"][:3]
            aff_d = [p for p in aff_roster if p["position"] == "D"][:10]
            aff_fwd = [p for p in aff_roster if p["position"] in ("C", "W")][:16]
            aff_chosen = aff_goalies[:max(2, len(aff_goalies))] + aff_d + aff_fwd

            if len(aff_goalies) < 2 or len(aff_d) < 5 or len(aff_fwd) < 9:
                # The game's loader tops every affiliate up to valid minimums
                # with synthesised depth fillers when still thin after overflow.
                print(f"  note AHL {aff_kw}: G{len(aff_goalies)} D{len(aff_d)} F{len(aff_fwd)} (incl. NHL overflow) — loader will fill the rest")

            clean_aff = resolve_faces(aff_chosen, faces_out, ahl_total, ahl_matched)

            affiliate_obj = {
                "city": aff_city,
                "nickname": aff_nick,
                "abbreviation": aff_abbr,
                "primary": aff_prim,
                "secondary": aff_sec,
                "players": clean_aff,
            }

        # Staff for this NHL club.
        raw_staff = nhl_staff.get(nick, [])
        clean_staff = resolve_staff_faces(raw_staff, faces_out, staff_total, staff_matched)

        team_obj = {
            "externalId": f"ehm-team-{abbr}",
            "city": city, "nickname": nick, "abbreviation": abbr,
            "primary": prim, "secondary": sec, "players": clean_nhl,
        }
        if affiliate_obj is not None:
            team_obj["affiliate"] = affiliate_obj
        if clean_staff:
            team_obj["staff"] = clean_staff
        meta = club_meta(nick)
        if meta and meta.get("arena"):
            team_obj["arena"] = meta["arena"]
            if meta.get("capacity"):
                team_obj["arenaCapacity"] = meta["capacity"]
        rn = club_retired(nick)
        if rn:
            team_obj["retiredNumbers"] = rn

        confs.setdefault(conf, {}).setdefault(div, []).append(team_obj)

    # ── Build the wider-world competitions (#95) ────────────────────────────
    import zlib
    MIN_COMP_ROSTER = 16  # need >= 2 G, 5 D, 9 F to ice quick-sim lines
    competitions_out = []
    comp_total, comp_matched = [0], [0]
    for league_proper, abbrev in COMP_LEAGUES.items():
        clubs_in = comp_clubs.get(league_proper)
        if not clubs_in:
            continue
        meta = comp_meta.get(league_proper.lower(), {})
        teams_out = []
        for club_name, bundle in sorted(clubs_in.items()):
            plist = bundle["players"]
            if len(plist) < MIN_COMP_ROSTER:
                continue
            info = bundle["info"]
            roster = sorted(plist, key=lambda p: -p["_ca"])
            g = [p for p in roster if p["position"] == "G"][:3]
            d = [p for p in roster if p["position"] == "D"][:8]
            fwd = [p for p in roster if p["position"] in ("C", "W")][:14]
            if len(g) < 2 or len(d) < 5 or len(fwd) < 9:
                continue
            chosen = g[:max(2, len(g))] + d + fwd
            clean = resolve_faces(chosen, faces_out, comp_total, comp_matched)
            prim, sec = COMP_PALETTE[zlib.crc32(club_name.encode("utf-8")) % len(COMP_PALETTE)]
            teams_out.append({
                "externalId": f"ehm-{abbrev.lower()}-{norm(club_name)}",
                "city": info["city"] or club_name,
                "nickname": info["nickname"],
                "abbreviation": info["abbr"],
                "primary": prim,
                "secondary": sec,
                "players": clean,
            })
        if len(teams_out) < 2:
            continue
        comp_obj = {
            "id": norm(league_proper),
            "name": league_proper,
            "abbrev": abbrev,
            "nation": meta.get("nation", ""),
            "level": meta.get("level", 1),
            "reputation": meta.get("reputation", 10),
            "teams": teams_out,
        }
        if meta.get("ageLimit", 0) > 0:
            comp_obj["upperAgeLimit"] = meta["ageLimit"]
        competitions_out.append(comp_obj)

    db = {
        "formatVersion": 1,
        "meta": {"name": "NHL (EHM import, dev)", "author": "local", "season": "2025-26"},
        "conferences": [
            {"name": cname, "divisions": [{"name": dname, "teams": dteams}
                                          for dname, dteams in divs.items()]}
            for cname, divs in confs.items()
        ],
        "competitions": competitions_out,
    }
    os.makedirs(out_dir, exist_ok=True)
    io.open(os.path.join(out_dir, "database.json"), "w", encoding="utf-8").write(
        json.dumps(db, ensure_ascii=False))
    io.open(os.path.join(out_dir, "faces.json"), "w", encoding="utf-8").write(
        json.dumps(faces_out, ensure_ascii=False))
    print(f"NHL teams: {sum(len(dd['teams']) for c in db['conferences'] for dd in c['divisions'])}")
    print(f"NHL players: {total[0]}  faces matched: {matched[0]} ({100*matched[0]//max(1,total[0])}%)")
    print(f"AHL players: {ahl_total[0]}  faces matched: {ahl_matched[0]} ({100*ahl_matched[0]//max(1,ahl_total[0])}%)")
    print(f"Staff: {staff_total[0]}  faces matched: {staff_matched[0]} ({100*staff_matched[0]//max(1,staff_total[0])}%)")
    print(f"Competitions: {len(competitions_out)}  "
          f"teams: {sum(len(c['teams']) for c in competitions_out)}  players: {comp_total[0]}")
    for c in competitions_out:
        print(f"  {c['abbrev']:6} {c['name'][:36]:36} teams={len(c['teams']):2}")

if __name__ == "__main__":
    main()
